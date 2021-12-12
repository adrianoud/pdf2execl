import openpyxl
import numpy as np
from sklearn import linear_model
from sklearn.metrics import mean_squared_error,r2_score
import matplotlib.pyplot as plt

print("汽侧流量（kg/s) + 进口压力(绝压Mpa) vs 汽侧压损(Mpa)")
wb = openpyxl.load_workbook("thermalDataFitting.xlsx")
for sh in wb.worksheets:
    x = []
    x1 = []
    x2 = []
    y = []
    y_pre = []
    y_error = []

    for i in range(29):
        x1.append(sh.cell(row=i + 3, column=12).value / 3600)  # 工质流量，单位转换为kg/s
        x2.append(sh.cell(row=i + 3, column=14).value + 0.1013)        # 进口压力，单位Mpa，转换成绝压
        y.append(sh.cell(row=i + 3, column=16).value)  # 压降，单位保持为Mpa

    x1 = np.array(x1)
    x2 = np.array(x2)

    x = np.array([x1, x1*x1, x2]).transpose()
    y = np.array(y).reshape(-1, 1)
    regr = linear_model.LinearRegression()
    regr.fit(x, y)
    y_pre = regr.predict(x)
    y_error = y_pre/y-1
    plt.scatter(x1,y)
    plt.savefig(sh.title)
    plt.close()
    if sh.title == "再热2" or sh.title == "再热1" or sh.title == "中过":
        print(sh.title)
        print(regr.coef_,regr.intercept_)
        print(r2_score(y,y_pre))
        print(max(y_error),min(y_error))
        print("\n")



    # pa = np.polyfit([x1,x2], y, 1)
    # print (pa)
