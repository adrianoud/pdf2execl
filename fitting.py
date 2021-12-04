import openpyxl
import numpy as np
import math

wb = openpyxl.load_workbook("thermalDataFitting.xlsx")


def tempdiff(a, b, c, d):
    return ((a - b) - (c - d)) / math.log((a - b) / (c - d))


for sh in wb.worksheets:
    x1_gasFlow = []
    x2_steamFlow = []
    y1_gasLoss = []
    y1_Pre = []
    y1_Pre_2 = []
    y1_error = []
    y1_error_2 = []
    y2_steamLoss = []
    y2_Pre = []
    y2_error = []
    y2_Pre_2 = []
    y2_error_2 = []
    y3_heatArea = []
    y3_Pre = []
    y3_error = []

    for i in range(29):
        x1_gasFlow.append(sh.cell(row=i + 3, column=2).value / 3.6)  # 单位转换为kg/s
        x2_steamFlow.append(sh.cell(row=i + 3, column=12).value / 3600)  # 单位转换为kg/s
        y1_gasLoss.append(sh.cell(row=i + 3, column=4).value)
        y2_steamLoss.append(sh.cell(row=i + 3, column=16).value * 1000)  # 单位转换为kpa
        y3_heatArea.append(sh.cell(row=i + 3, column=20).value / tempdiff(sh.cell(row=i + 3, column=5).value, sh.cell(row=i + 3, column=17).value,
        sh.cell(row=i + 3, column=6).value, sh.cell(row=i + 3, column=18).value))

    pa_x1y1 = np.polyfit(x1_gasFlow, y1_gasLoss, 1)
    pa_x1y1_2 = np.polyfit(x1_gasFlow, y1_gasLoss, 2)

    pa_x2y2 = np.polyfit(x2_steamFlow, y2_steamLoss, 1)
    pa_x2y2_2 = np.polyfit(x2_steamFlow, y2_steamLoss, 2)

    pa_x2y3 = np.polyfit(x2_steamFlow, y3_heatArea, 1)

    for j in range(29):
        y1_Pre.append(pa_x1y1[0] * x1_gasFlow[j] + pa_x1y1[1])
        y1_error.append(1 - y1_Pre[j] / y1_gasLoss[j])
        y1_Pre_2.append(pa_x1y1_2[0] * x1_gasFlow[j] ** 2 + pa_x1y1_2[1] * x1_gasFlow[j] + pa_x1y1_2[2])
        y1_error_2.append(1 - y1_Pre_2[j] / y1_gasLoss[j])

        y2_Pre.append(pa_x2y2[0] * x2_steamFlow[j] + pa_x2y2[1])
        y2_error.append(1 - y2_Pre[j] / y2_steamLoss[j])
        y2_Pre_2.append(pa_x2y2_2[0] * x2_steamFlow[j] ** 2 + pa_x2y2_2[1] * x2_steamFlow[j] + pa_x2y2_2[2])
        y2_error_2.append(1 - y2_Pre_2[j] / y2_steamLoss[j])

        y3_Pre.append(pa_x2y3[0] * x2_steamFlow[j] + pa_x2y3[1])
        y3_error.append(1 - y3_Pre[j] / y3_heatArea[j])


    x1y1_Square = np.corrcoef(y1_gasLoss, y1_Pre)[0, 1]
    x1y1_Square_2 = np.corrcoef(y1_gasLoss, y1_Pre_2)[0, 1]

    x2y2_Square = np.corrcoef(y2_steamLoss, y2_Pre)[0, 1]
    x2y2_Square_2 = np.corrcoef(y2_steamLoss, y2_Pre_2)[0, 1]

    x2y3_Square = np.corrcoef(y2_steamLoss, y3_Pre)[0, 1]

    print("======================================", sh.title, "=======================================")
    print('烟侧流量 vs 烟侧压损')
    print('线性回归：')
    print("系数：", pa_x1y1, "R平方：", x1y1_Square, "最大正误差：", max(y1_error), "最大负误差：", min(y1_error))
    print('二次回归：')
    print("系数：", pa_x1y1_2, "R平方：", x1y1_Square_2, "最大正误差：", max(y1_error_2), "最大负误差：", min(y1_error_2))
    if (x1y1_Square < 0.9 or max(y1_error) > 0.1 or min(y1_error) < -0.1) and (x1y1_Square_2 < 0.9 or max(y1_error_2) > 0.1 or min(y1_error_2) < -0.1):
        print("!!!!!!&&&&!!!!!!!!!warning!!!!!!!!!&&&!!!!!!!!")
    print("\n")

    print('汽侧流量 vs 汽侧压损')
    print('线性回归：')
    print("系数：", pa_x2y2, "R平方：", x2y2_Square, "最大正误差：", max(y2_error), "最大负误差：", min(y2_error))
    print('二次回归：')
    print("系数：", pa_x2y2_2, "R平方：", x2y2_Square_2, "最大正误差：", max(y2_error_2), "最大负误差：", min(y2_error_2))
    if (x2y2_Square < 0.9 or max(y2_error) > 0.1 or min(y2_error) < -0.1) and (x2y2_Square_2 < 0.9 or max(y2_error_2) > 0.1 or min(y2_error_2) < -0.1):
        print("!!!!!!&&&&!!!!!!!!!warning!!!!!!!!!&&&!!!!!!!!")
    print("\n")
    print('汽侧流量 vs 当量换热面积, 线性回归')
    print("系数：", pa_x2y3, "R平方：", x2y3_Square, "最大正误差：", max(y3_error), "最大负误差：", min(y3_error))
    if x2y3_Square < 0.9 or max(y3_error) > 0.1 or min(y3_error) < -0.1:
        print("!!!!!!&&&&!!!!!!!!!warning!!!!!!!!!&&&!!!!!!!!")
    print("\n")
#   print(x1_gasFlow)
#   print(y1_gasLoss)
#   print(y1_Pre)
#   print(y1_error)
