import openpyxl
import numpy as np
from fitting import tempdiff
from sklearn import linear_model
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook("thermalDataFitting.xlsx")
np.set_printoptions(precision=30, suppress=True)
for sh in wb.worksheets:
# sh = wb["高过2"]
    x = []
    x1 = []
    x2 = []
    # x3 = []
    y = []
    y_pre = []
    y_error = []

    for i in range(29):
        x1.append(sh.cell(row=i + 3, column=2).value / 3.6)  # 烟气流量，单位转换为kg/s
        x2.append(sh.cell(row=i + 3, column=12).value / 3600)  # 蒸汽流量，单位转换为kg/s
        y.append(sh.cell(row=i + 3, column=20).value * 1000 / 3.6 / tempdiff(sh.cell(row=i + 3, column=5).value,
                                                                             sh.cell(row=i + 3, column=17).value,
                                                                             sh.cell(row=i + 3, column=6).value,
                                                                             sh.cell(row=i + 3, column=18).value))
    x1 = np.array(x1)
    x2 = np.array(x2)
    # x3 = np.array(x3)
    x = np.array([x1, x2]).transpose()
    # print(x,y)
    y = np.array(y).reshape(-1, 1)
    regr = linear_model.LinearRegression()
    regr.fit(x, y)
    y_pre = regr.predict(x)
    y_error = y_pre / y - 1
    plt.scatter(x1, y)
    plt.savefig(sh.title)
    plt.close()
    print(sh.title)
    print(regr.coef_, regr.intercept_)
    print(r2_score(y, y_pre))
    # print(y_error)
    print(max(y_error), min(y_error))
    print("\n")

