import openpyxl
import numpy as np
from sklearn import linear_model
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook("thermalDataFitting.xlsx")
np.set_printoptions(precision=30, suppress=True)
for sh in wb.worksheets:
# sh = wb["高过2"]
    x = []
    x1 = []
    x2 = []
    x3 = []
    y = []
    y_pre = []
    y_error = []

    for i in range(29):
        x1.append(sh.cell(row=i + 3, column=2).value / 3.6)  # 烟气流量，单位转换为kg/s
        x2.append(sh.cell(row=i + 3, column=5).value + 273.15)  # 烟气进口温度，单位转换为K
        # x3.append(sh.cell(row=i + 3, column=3).value * 1000000 + 0.10135)  # 烟气进口压力，单位转换为Mpa.a
        y.append(sh.cell(row=i + 3, column=4).value / 1000000)  # 烟气压损， 单位转换为Mpa

    x1 = np.array(x1)
    x2 = np.array(x2)
    # x3 = np.array(x3)
    x = np.array([x1*x1, x1, x2]).transpose()
    # print(x,y)
    y = np.array(y).reshape(-1, 1)
    regr = linear_model.LinearRegression()
    regr.fit(x, y)
    y_pre = regr.predict(x)
    y_error = y_pre / y - 1
    plt.scatter(x1, y)
    plt.savefig(sh.title)
    plt.close()
    print(sh.title)
    print(regr.coef_, regr.intercept_)
    print(r2_score(y, y_pre))
    # print(y_error)
    print(max(y_error), min(y_error))
    print("\n")

