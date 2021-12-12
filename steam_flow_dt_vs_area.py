import openpyxl
import numpy as np
from sklearn import linear_model
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt
from fitting import tempdiff

wb = openpyxl.load_workbook("thermalDataFitting.xlsx")
for sh in wb.worksheets:
    x = []
    x1 = []
    x2 = []
    y = []
    y_pre = []
    y_error = []

    for i in range(29):
        x1.append(sh.cell(row=i + 3, column=12).value / 3600)  # 工质流量，单位转换为kg/s
        x2.append(sh.cell(row=i + 3, column=19).value)  # 工质温差
        # 换热量单位转为kw
        y.append(sh.cell(row=i + 3, column=20).value * 1000 / 3.6 / tempdiff(sh.cell(row=i + 3, column=5).value,
                                                                             sh.cell(row=i + 3, column=17).value,
                                                                             sh.cell(row=i + 3, column=6).value,
                                                                             sh.cell(row=i + 3, column=18).value))
    x = np.array([x1, x2]).transpose()
    y = np.array(y).reshape(-1, 1)
    regr = linear_model.LinearRegression()
    regr.fit(x, y)
    y_pre = regr.predict(x)
    y_error = y_pre / y - 1
    plt.scatter(x1, y)
    plt.savefig(sh.title)
    plt.close()
    print(sh.title)
    print(regr.coef_, regr.intercept_)
    print(r2_score(y, y_pre))
    print(max(y_error), min(y_error))
    print("\n")

    # pa = np.polyfit([x1,x2], y, 1)
    # print (pa)
