import pdfplumber
import openpyxl
import numpy as np
wb = openpyxl.load_workbook("thermalData.xlsx")
pdf = pdfplumber.open("111.pdf")

for m in range(8):
    for k in range(29):
        cells = []
        page0 = pdf.pages[2*k+2]
        table = page0.extract_text()
        rows = table.split("\n")
        for i in range(len(rows)):
            cells.append([])
            cols = rows[i].split(" ")
            for j in range(len(cols)):
                cells[i].append(cols[j])
#        print(cells[0][0])

        if cells[1][m+3] in wb.sheetnames:
            sheet = wb[cells[1][m+3]]
        else:
            sheet = wb.create_sheet(cells[1][m+3])

        sheet.cell(2, 1).value = "工况："
        sheet.cell(k+3, 1).value = cells[0][0]
        sheet.cell(1, 2).value = "烟气侧"
        sheet.cell(1, 12).value = "工质侧"
        for i in range(2, 8):
            sheet.cell(2, i).value = cells[i+1][0] + ' ' + "(" + cells[i+1][1] + ")"
            sheet.cell(k+3, i).value = float(cells[i + 1][m+3])

        for i in range(8, 12):
            sheet.cell(2, i).value = cells[i+1][0] + ' ' + "(" + cells[i+1][1] + ")"
            sheet.cell(k+3, i).value = float(cells[i + 1][m+2])

        for i in range(12, 22):
            sheet.cell(2, i).value = cells[i+4][0] + ' ' + "(" + cells[i+4][1] + ")"
            sheet.cell(k+3, i).value = float(cells[i+4][m+2])

        wb.save("thermalData.xlsx")

m = 0
k = 0
i = 0
j = 0

for m in range(7):
    for k in range(29):
        cells = []
        page = pdf.pages[2*k+3]
        table = page.extract_text()
        rows = table.split("\n")
        for i in range(len(rows)):
            cells.append([])
            cols = rows[i].split(" ")
            for j in range(len(cols)):
                cells[i].append(cols[j])
        print(cells[0][0])

        if cells[1][m+2] in wb.sheetnames:
            sheet = wb[cells[1][m+2]]
        else:
            sheet = wb.create_sheet(cells[1][m+2])

        sheet.cell(2, 1).value = "工况："
        sheet.cell(k+3, 1).value = cells[0][0]
        sheet.cell(1, 2).value = "烟气侧"
        sheet.cell(1, 12).value = "工质侧"
        for i in range(2, 12):
            sheet.cell(2, i).value = cells[i+1][0] + ' ' + "(" + cells[i+1][1] + ")"
            sheet.cell(k+3, i).value = float(cells[i + 1][m+2])

        for i in range(12, 22):
            sheet.cell(2, i).value = cells[i+4][0] + ' ' + "(" + cells[i+4][1] + ")"
            sheet.cell(k+3, i).value = float(cells[i + 4][m+2])

        wb.save("thermalData.xlsx")